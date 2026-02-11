from django.shortcuts import render
from django.core.paginator import Paginator
from django.views import generic
from django.db.models import Avg, Max, Min, Count, Sum
from datetime import datetime
from datetime import timedelta
from django.utils.timezone import now
from django import forms
from .forms import InputForm,  CustomUserCreationForm, EmailUpdateForm, UpdateNameForm, CustomPasswordChangeForm, DeepSeekConfigForm
from django.utils.decorators import method_decorator
from django.views.generic.detail import DetailView
from django.contrib.auth.decorators import permission_required
from django.core.cache import cache
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required

from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.contrib.auth.models import Group
from django.contrib.auth import authenticate, login
from django.shortcuts import get_object_or_404
from django.http import Http404
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, JsonResponse
from django.http import HttpResponse
from django.urls import reverse
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.db import transaction
from django.db.models import Q
from django.views.decorators.http import require_http_methods
import json
import csv
from time import sleep
import logging
import re
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
import json
from django.template.loader import render_to_string
from .translate import baidu_translate, parse_json_to_string, check_api_keys
from .deepseek import deepseek_translate, check_deepseek_keys, call_deepseek_api
import openpyxl
from django.db import transaction
from .models import Item, Category, Proficiency, ReviewDay, DeepSeekConfig
import difflib
import uuid
from .utils import fetch_and_merge_translation
import markdown
from django.conf import settings
import os

logger = logging.getLogger(__name__)

# Create your views here.
from .models import Item, Proficiency, Category, ReviewDay

def custom_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # 使用 authenticate 进行身份验证
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # 如果用户验证通过，则登录并重定向
            login(request, user)
            return redirect('home')  # 登录成功后可以重定向到首页或其他页面
        else:
            # 如果用户名或密码错误，使用消息框架显示错误信息
            messages.error(request, "用户名或密码无效，请检查后重试。")

    return render(request, 'registration/login.html')


def register(request):
    if request.method == 'POST':
        random_id = request.POST.get('random_id', None)  # 获取随机 ID
        # 重构 POST 数据，将动态字段映射回标准字段
        if random_id:
            mapped_post = {
                'username': request.POST.get(f'random_username_{random_id}', ''),
                'email': request.POST.get('email', ''),
                'first_name': request.POST.get('first_name', ''),
                'last_name': request.POST.get('last_name', ''),
                'password1': request.POST.get(f'random_password1_{random_id}', ''),
                'password2': request.POST.get(f'random_password2_{random_id}', ''),
            }
            form = CustomUserCreationForm(mapped_post)
        else:
            form = CustomUserCreationForm(request.POST)

        if form.is_valid():
            try:
                user = form.save()
                username = form.cleaned_data.get('username')

                # 加入 Public 组
                public_group, created = Group.objects.get_or_create(name='Public')
                user.groups.add(public_group)
                user.is_staff = True
                user.save()

                # 创建默认类别和复习计划
                Category.objects.create(user=user, name="单词", sort_order=1, is_default=True)
                review_days = [1, 2, 4, 7, 15, 30, 90, 180, 365]
                ReviewDay.objects.bulk_create(
                    [ReviewDay(user=user, day=day) for day in review_days]
                )

                messages.success(request, f'Account {username} created successfully!')
                return redirect('login')

            except Exception as e:
                logger.error(f"Error during registration: {e}")
                messages.error(request, f"Registration failed: {e}")
        else:
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    if field == 'password2':
                        error = error.replace('password2', 'Password confirmation')
                    error_messages.append(f"<p>{error}</p>")

            form_errors = "".join(error_messages)
            logger.warning(f"Form validation failed: {form.errors}")
            messages.error(request, f"<p>Please fix the following errors: {form_errors}</p>")
    else:
        random_id = uuid.uuid4().hex  # 生成一个随机 ID
        form = CustomUserCreationForm()
    
    return render(request, 'registration/register.html', {'form': form, 'random_id': random_id})


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'

@login_required
def user_profile(request):
    if request.method == 'POST':
        # Update email
        if 'update_email' in request.POST:
            email_form = EmailUpdateForm(request.user, request.POST)
            if email_form.is_valid():
                request.user.email = email_form.cleaned_data.get('email')
                request.user.save()
                messages.success(request, "Email updated successfully.")
            else:
                messages.error(request, "Failed to update email. Please check the errors.")

        # Update name
        elif 'update_profile' in request.POST:
            name_form = UpdateNameForm(request.POST, instance=request.user)
            if name_form.is_valid():
                name_form.save()
                messages.success(request, "Name updated successfully.")
            else:
                messages.error(request, "Failed to update name. Please check the errors.")

        # Update password
        elif 'change_password' in request.POST:
            password_form = CustomPasswordChangeForm(user=request.user, data=request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)  # Keep user logged in
                messages.success(request, "Password updated successfully.")
            else:
                messages.error(request, "Failed to update password. Please check the errors.")

        return redirect('user_profile')

    else:
        email_form = EmailUpdateForm(request.user)
        name_form = UpdateNameForm(instance=request.user)
        password_form = CustomPasswordChangeForm(user=request.user)

    return render(request, 'user_profile.html', {
        'email_form': email_form,
        'name_form': name_form,
        'password_form': password_form,
    })


def home(request):
    # 检查用户是否登录
    if request.user.is_authenticated:
        # 获取当前用户的所有 items
        items = Item.objects.filter(user=request.user)
        # 统计数据
        total_items = items.count()
        if total_items > 0:
            first_item_date = items.order_by('inputDate').first().inputDate
            days_since_first_item = (now().date() - first_item_date).days
        else:
            days_since_first_item = 0
        
        # 判断如何显示用户名
        if request.user.first_name and request.user.last_name:
            display_name = f"{request.user.first_name} {request.user.last_name}"  # 合并 first_name 和 last_name
        elif request.user.first_name:
            display_name = request.user.first_name  # 只有 first_name
        elif request.user.last_name:
            display_name = request.user.last_name  # 只有 last_name
        else:
            display_name = request.user.username  # 都没有，使用 username
        
        context = {
            'display_name': display_name,  # 修改为 display_name
            'total_items': total_items,
            'days_since_first_item': days_since_first_item,
        }
        # 用户已登录，返回登录后的首页
        return render(request, 'home_logged_in.html', context)
    else:
        # 用户未登录，返回未登录的首页
        return render(request, 'home_logged_out.html')


# @login_required
# def index(request):
#     """
#     Index view to show the total count of items and the count for each category.
#     """
#     # 获取当前登录用户
#     user = request.user

#     # 查询用户的所有条目和分类
#     total_items = Item.objects.filter(user=user).count()  # 总条目数
#     categories = Category.objects.filter(user=user)  # 当前用户的所有分类
#     category_counts = {
#         category.name: Item.objects.filter(user=user, category=category).count()
#         for category in categories
#     }

#     # 渲染上下文
#     context = {
#         'total_items': total_items,
#         'category_counts': category_counts,
#     }

#     return render(request, 'index.html', context)

@login_required
@ensure_csrf_cookie
def item_list(request):
    # 获取筛选的分类ID
    category_id = request.GET.get('category')

    # 获取当前登录用户的所有 Item
    item_list = Item.objects.filter(user=request.user)

    # 如果指定了分类，进行筛选
    if category_id:
        item_list = item_list.filter(category_id=category_id)

    item_list = item_list.order_by('-inputDate')

    # 获取所有分类（包括没有单词的分类）
    all_categories = Category.objects.filter(user=request.user).order_by('sort_order', 'name')

    # 统计每个类别下的条目数量
    category_stats = []
    for category in all_categories:
        count = Item.objects.filter(user=request.user, category=category).count()
        category_stats.append({
            'category__name': category.name,
            'category__id': category.id,
            'count': count
        })

    # 按 count 降序排序
    category_stats.sort(key=lambda x: x['count'], reverse=True)

    # 确保 item_list 不为空时才进行分页
    if item_list.exists():
        paginator = Paginator(item_list, 50)  # 每页 50 个
        page_number = request.GET.get('page')  # 获取当前页码
        page_obj = paginator.get_page(page_number)
    else:
        # 如果 item_list 为空，设置 page_obj 为一个空列表或自定义的对象
        page_obj = []

    context = {
        'page_obj': page_obj,
        'category_stats': category_stats,
        'selected_category': int(category_id) if category_id else None
    }

    # 如果是 AJAX 请求，返回 JSON 数据
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        items_data = []
        for item in page_obj if page_obj else []:
            items_data.append({
                'id': item.id,
                'item': item.item,
                'category': item.category.name,
                'category_id': item.category.id,
                'inputDate': item.inputDate.isoformat() if item.inputDate else None,
                'next_review_date': item.next_review_date.isoformat() if item.next_review_date else None,
                'detail_url': reverse('item-detail', args=[item.id])
            })
        return JsonResponse({
            'items': items_data,
            'has_next': page_obj.has_next() if page_obj else False,
            'has_previous': page_obj.has_previous() if page_obj else False,
            'current_page': page_obj.number if page_obj else 1,
            'total_pages': page_obj.paginator.num_pages if page_obj else 1
        })

    # 渲染模板，传递分页对象和类别统计信息
    return render(request, 'list.html', context)


# ==================== 分类管理视图 ====================

@login_required
@ensure_csrf_cookie
def category_list(request):
    """返回当前用户的所有分类（JSON）"""
    categories = Category.objects.filter(user=request.user).order_by('sort_order', 'name')
    data = [{
        'id': cat.id,
        'name': cat.name,
        'sort_order': cat.sort_order,
        'is_default': cat.is_default,
        'item_count': Item.objects.filter(user=request.user, category=cat).count()
    } for cat in categories]
    return JsonResponse({'categories': data})


@login_required
@ensure_csrf_cookie
def category_create(request):
    """创建新分类"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            name = data.get('name', '').strip()
            if not name:
                return JsonResponse({'success': False, 'error': '分类名称不能为空'})

            # 检查是否已存在
            if Category.objects.filter(user=request.user, name=name).exists():
                return JsonResponse({'success': False, 'error': '该分类已存在'})

            # 获取当前最大 sort_order
            max_order = Category.objects.filter(user=request.user).aggregate(
                max_order=Max('sort_order')
            )['max_order'] or 0

            category = Category.objects.create(
                user=request.user,
                name=name,
                sort_order=max_order + 1
            )
            return JsonResponse({'success': True, 'category': {'id': category.id, 'name': category.name}})
        except Exception as e:
            logger.error(f"Error creating category: {e}")
            return JsonResponse({'success': False, 'error': str(e)})


@login_required
@ensure_csrf_cookie
def category_update(request, category_id):
    """更新分类名称"""
    if request.method == 'POST':
        try:
            category = get_object_or_404(Category, id=category_id, user=request.user)

            data = json.loads(request.body)
            name = data.get('name', '').strip()
            if not name:
                return JsonResponse({'success': False, 'error': '分类名称不能为空'})

            # 检查新名称是否与其他分类重复
            if Category.objects.filter(user=request.user, name=name).exclude(id=category_id).exists():
                return JsonResponse({'success': False, 'error': '该分类名称已存在'})

            category.name = name
            category.save()
            return JsonResponse({'success': True})
        except Exception as e:
            logger.error(f"Error updating category: {e}")
            return JsonResponse({'success': False, 'error': str(e)})


@login_required
@ensure_csrf_cookie
def category_delete(request, category_id):
    """删除分类（级联删除该分类下的所有单词）"""
    if request.method == 'POST':
        try:
            category = get_object_or_404(Category, id=category_id, user=request.user)

            # 统计该分类下的单词数量
            item_count = Item.objects.filter(user=request.user, category=category).count()

            # 删除分类（由于设置了 CASCADE，会自动删除该分类下的所有单词）
            category.delete()

            return JsonResponse({
                'success': True,
                'deleted_items': item_count,
                'message': f'已删除分类及其 {item_count} 个单词'
            })
        except Exception as e:
            logger.error(f"Error deleting category: {e}")
            return JsonResponse({'success': False, 'error': str(e)})


# ==================== 批量操作视图 ====================

@login_required
@ensure_csrf_cookie
def batch_delete_items(request):
    """批量删除单词"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            item_ids = data.get('item_ids', [])

            if not item_ids:
                return JsonResponse({'success': False, 'error': '请选择要删除的单词'})

            # 删除属于当前用户的单词
            count, _ = Item.objects.filter(user=request.user, id__in=item_ids).delete()

            return JsonResponse({'success': True, 'deleted_count': count})
        except Exception as e:
            logger.error(f"Error batch deleting items: {e}")
            return JsonResponse({'success': False, 'error': str(e)})


@login_required
@ensure_csrf_cookie
def batch_move_items(request):
    """批量移动单词到指定分类"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            item_ids = data.get('item_ids', [])
            target_category_id = data.get('target_category_id')

            if not item_ids:
                return JsonResponse({'success': False, 'error': '请选择要移动的单词'})

            if not target_category_id:
                return JsonResponse({'success': False, 'error': '请选择目标分类'})

            # 验证目标分类属于当前用户
            target_category = get_object_or_404(Category, id=target_category_id, user=request.user)

            # 更新单词的分类
            count = Item.objects.filter(
                user=request.user,
                id__in=item_ids
            ).update(category=target_category)

            return JsonResponse({'success': True, 'moved_count': count})
        except Exception as e:
            logger.error(f"Error batch moving items: {e}")
            return JsonResponse({'success': False, 'error': str(e)})


@method_decorator(login_required, name='dispatch')  # 确保用户已登录
class ItemDetailView(DetailView):
    model = Item
    template_name = 'item_detail.html'  # 设置模板路径

    def get_queryset(self):
        # 只允许当前登录用户访问属于自己的 Item
        return Item.objects.filter(user=self.request.user)

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        # 验证是否属于当前用户
        if obj.user != self.request.user:
            # 返回自定义提示页面
            return render(self.request, 'EAW/item_not_found.html', status=404)
        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        item = context.get('object')
        from datetime import date, timedelta

        # 与其他地方保持一致的间隔定义
        intervals = [0, 1, 2, 4, 7, 15, 30, 90, 180]

        # 基准日期：优先使用 initDate，否则使用 inputDate，若都无则使用今天
        base_date = item.initDate or item.inputDate or date.today()

        # unfamiliar_history 可能为 None
        uh = item.unfamiliar_history or []

        schedule = []
        today = date.today()
        # 计算当前的索引（如果 current_interval 存的是天数）
        try:
            current_index = intervals.index(item.current_interval) if item.current_interval is not None else None
        except ValueError:
            current_index = None

        for idx, days in enumerate(intervals):
            scheduled_date = None
            try:
                scheduled_date = base_date + timedelta(days=days)
            except Exception:
                scheduled_date = None

            # 是否已完成：根据 current_index（interval 的索引）判断
            completed = False
            if current_index is not None and current_index > idx:
                completed = True
            else:
                # 如果在 unfamiliar_history 中存在针对该间隔、且日期等于计划日期的记录（用户已点评，无论 YES/NO），视为已完成
                try:
                    if scheduled_date and any(((r.get('interval') == days or r.get('interval') == idx) and r.get('date') == scheduled_date.isoformat()) for r in uh):
                        completed = True
                except Exception:
                    pass

            # 是否逾期（已过计划日但未完成）
            overdue = False
            if scheduled_date and scheduled_date < today and not completed:
                overdue = True

            # 熟悉判断：如果 completed 并且 unfamiliar_history 中没有该 interval 的记录，则视为熟悉
            unfamiliar_records = [r for r in uh if r.get('interval') == days or r.get('interval') == idx]
            if completed:
                familiar = (len(unfamiliar_records) == 0)
            else:
                familiar = None

            schedule.append({
                'index': idx,
                'days': days,
                'scheduled_date': scheduled_date,
                'completed': completed,
                'overdue': overdue,
                'familiar': familiar,
                'unfamiliar_records': unfamiliar_records,
            })

        context['review_schedule'] = schedule
        return context


@method_decorator(login_required, name='dispatch')
class ItemUpdateView(generic.UpdateView):
    model = Item
    template_name = 'item_form.html'
    fields = ['item', 'content', 'category', 'uk_phonetic', 'us_phonetic', 'src_tts', 'inputDate', 'initDate', 'proficiency']
    
    def get_queryset(self):
        # 只允许用户编辑自己的Item
        return Item.objects.filter(user=self.request.user)
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # 限制category只显示当前用户的类别
        form.fields['category'].queryset = Category.objects.filter(user=self.request.user)
        # 添加Bootstrap样式
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ != 'CheckboxInput':
                field.widget.attrs['class'] = 'form-control'
        return form
    
    def form_valid(self, form):
        messages.success(self.request, f'单词 "{form.instance.item}" 已成功更新！')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('item-list')


@method_decorator(login_required, name='dispatch')
class ItemDeleteView(generic.DeleteView):
    model = Item
    template_name = 'item_confirm_delete.html'
    
    def get_queryset(self):
        # 只允许用户删除自己的Item
        return Item.objects.filter(user=self.request.user)
    
    def delete(self, request, *args, **kwargs):
        item = self.get_object()
        messages.success(request, f'单词 "{item.item}" 已成功删除！')
        return super().delete(request, *args, **kwargs)
    
    def get_success_url(self):
        return reverse('item-list')


@login_required
def SearchView(request):
    word = ''
    query = Item.objects.none()

    # 获取搜索关键词
    search_input = request.GET.get('q', '').strip()
    logger.debug(f"Search input received: {search_input}")

    if search_input == '':  # 如果没有输入关键词
        if 'q' in request.GET:
            word = 'No search input.'
        return render(
            request,
            'search.html',
            context={'word': word},
        )
    else:
        # 在当前用户的数据库中搜索 item 字段包含搜索关键词的条目
        try:
            query = Item.objects.filter(user=request.user, item__icontains=search_input)
            if not query.exists():
                word = 'No search result.'
        except Exception as e:
            logger.error(f"Error during search query: {e}")
            return JsonResponse({'error': 'Error processing your search query'}, status=500)

    # 如果是 AJAX 请求，返回 JSON 数据
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            # 渲染搜索结果的 HTML
            html = render_to_string('search_results.html', {'query': query, 'word': word})
            logger.debug(f"Generated HTML for search results.")
            return JsonResponse({'html': html})
        except Exception as e:
            logger.error(f'Error during search result rendering: {e}')
            return JsonResponse({'error': 'Failed to generate results.'}, status=500)

    # 如果不是 AJAX 请求，返回正常的 HTML 页面
    return render(
        request,
        'search_results.html',
        context={'query': query, 'word': word},
    )

#复习单词的功能
@login_required
def ReviewHomeView(request):
    # print("Request routed to ReviewHomeView")  # 调试
    today = datetime.today().date()
    #print(today)
    return render(
        request,
        'review_home.html',
        context={'today': today}
    )


# Calendar views and APIs (FullCalendar integration)
@login_required
@login_required
def calendar_month_view(request):
    """渲染日历页面（前端使用 FullCalendar 拉取事件）。"""
    return render(request, 'calendar_month.html', {
        'today': datetime.today().date()
    })


@login_required
def calendar_events_api(request):
    """返回指定时间范围内按天聚合的复习统计，FullCalendar 兼容格式。"""
    try:
        start = request.GET.get('start')
        end = request.GET.get('end')
        
        # 添加日志
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Calendar API called: start={start}, end={end}, user={request.user}")
        
        def parse_date_safe(s, default):
            if not s:
                return default
            try:
                return datetime.fromisoformat(s).date()
            except Exception:
                try:
                    # fallback: take first 10 chars (YYYY-MM-DD)
                    return datetime.fromisoformat(s[:10]).date()
                except Exception:
                    return default

        start_date = parse_date_safe(start, datetime.today().date() - timedelta(days=30))
        end_date = parse_date_safe(end, datetime.today().date() + timedelta(days=30))
        
        logger.info(f"Parsed dates: start_date={start_date}, end_date={end_date}")

        days = (end_date - start_date).days
        events = []

        from django.db.models import Q
        
        for i in range(days + 1):
            d = start_date + timedelta(days=i)
            
            # 分别查询：今日计划、额外复习（不再统计逾期）
            # 1. 今日计划：next_review_date 正好是今天
            today_planned = Item.objects.filter(
                user=request.user,
                next_review_date=d
            )
            
            # 2. 额外复习：需要额外复习且在复习期内
            extra_review = Item.objects.filter(
                user=request.user
            ).filter(
                Q(needs_extra_review=True, extra_review_since__lte=d, next_review_date__gt=d) |
                Q(next_review_date__isnull=True, needs_extra_review=True, extra_review_since__lte=d)
            )

            # 统计今日计划的完成情况
            today_total = today_planned.count()
            today_completed = 0
            today_pending = 0
            
            for it in today_planned:
                reviewed = False
                try:
                    uh = it.unfamiliar_history or []
                    for r in uh:
                        if r.get('date') == d.isoformat():
                            reviewed = True
                            break
                except Exception:
                    pass
                
                if reviewed:
                    today_completed += 1
                else:
                    today_pending += 1
            
            # 统计额外复习的
            extra_count = extra_review.count()
            
            # 只有当天有计划或额外复习时才显示
            if today_total == 0 and extra_count == 0:
                continue

            # 构建标题：主要显示今日计划
            title_parts = []
            if today_pending > 0:
                title_parts.append(f"今日: {today_pending}待")
            if today_completed > 0:
                title_parts.append(f"{today_completed}已完成")
            if extra_count > 0:
                title_parts.append(f"额外: {extra_count}")
            
            title = " · ".join(title_parts) if title_parts else "无待办"

            # 确定事件颜色
            if today_pending == 0 and extra_count == 0:
                color = '#28a745'  # 绿色：全部完成
            elif today_pending > 0:
                color = '#ffc107'  # 黄色：有今日待办
            else:
                color = '#007bff'  # 蓝色：其他情况
            
            events.append({
                'id': d.isoformat(),
                'title': title,
                'start': d.isoformat(),
                'allDay': True,
                'backgroundColor': color,
                'borderColor': color,
                'extendedProps': {
                    'today_planned': today_total,
                    'today_pending': today_pending,
                    'today_completed': today_completed,
                    'extra': extra_count
                }
            })

        logger.info(f"Returning {len(events)} events")
        return JsonResponse(events, safe=False)

    except Exception as e:
        import traceback
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Calendar API error: {str(e)}")
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def calendar_day_items_api(request):
    """返回指定日期的复习条目列表（JSON）。"""
    try:
        date_str = request.GET.get('date')
        if not date_str:
            return JsonResponse({'error': '缺少日期参数'}, status=400)
        
        try:
            d = datetime.fromisoformat(date_str).date()
        except Exception:
            try:
                d = datetime.fromisoformat(date_str[:10]).date()
            except Exception:
                return JsonResponse({'error': '无效的日期格式'}, status=400)

        from django.db.models import Q
        
        # 只查询今日计划和额外复习，不包括逾期
        review_items = Item.objects.filter(user=request.user).filter(
            Q(next_review_date=d) |
            Q(needs_extra_review=True, extra_review_since__lte=d, next_review_date__gt=d) |
            Q(next_review_date__isnull=True, needs_extra_review=True, extra_review_since__lte=d)
        ).order_by('next_review_date', 'item')

        items = []
        intervals = [0, 1, 2, 4, 7, 15, 30, 90, 180]
        
        for it in review_items:
            try:
                uh = it.unfamiliar_history or []
                reviewed_today = any(r.get('date') == d.isoformat() for r in uh)
                
                # 统计当前周期不熟悉次数
                if it.current_interval in intervals:
                    cur_day = it.current_interval
                    try:
                        cur_idx = intervals.index(cur_day)
                    except Exception:
                        cur_idx = None
                else:
                    try:
                        cur_idx = int(it.current_interval)
                        cur_day = intervals[cur_idx] if 0 <= cur_idx < len(intervals) else it.current_interval
                    except Exception:
                        cur_idx = None
                        cur_day = it.current_interval

                unfamiliar_count = sum(
                    1 for r in uh 
                    if r.get('interval') == cur_day or r.get('interval') == cur_idx
                )
                
            except Exception:
                reviewed_today = False
                unfamiliar_count = 0
            
            # 判断是否为额外复习
            is_extra = (
                it.needs_extra_review and 
                it.extra_review_since and 
                it.extra_review_since <= d and 
                (not it.next_review_date or it.next_review_date > d)
            )

            items.append({
                'id': it.id,
                'item': it.item,
                'interval_day': it.current_interval,
                'next_review_date': it.next_review_date.isoformat() if it.next_review_date else None,
                'unfamiliar_count': unfamiliar_count,
                'reviewed_today': reviewed_today,
                'is_extra_review': is_extra,
                'detail_url': reverse('item-detail', args=[it.pk])
            })

        return JsonResponse({
            'date': d.isoformat(),
            'total': len(items),
            'items': items
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def ReviewView(request, year, month, day):
    from django.db.models import Q
    
    # 创建选择的复习日期
    d1 = f"{year}-{month}-{day}"
    reviewDate = datetime.strptime(d1, '%Y-%m-%d').date()
    
    # 如果是POST请求，处理用户选择的日期
    if request.method == 'POST':
        review_date_str = request.POST.get('review_date')
        if review_date_str:
            reviewDate = datetime.strptime(review_date_str, '%Y-%m-%d').date()
    
    # 获取URL参数
    show_mastered = request.GET.get('show_mastered', 'false').lower() == 'true'
    per_page = int(request.GET.get('per_page', 10))
    page_number = request.GET.get('page', 1)
    
    # 查询需要复习的单词
    review_items_query = Item.objects.filter(user=request.user).filter(
        Q(next_review_date__lte=reviewDate) |  # 正式复习日到了
        Q(needs_extra_review=True, extra_review_since__lte=reviewDate, next_review_date__gt=reviewDate)  # 额外复习期内且未到正式复习日
    )
    
    # 收集复习items信息
    # 注意：只要next_review_date到了就应该复习，不管proficiency是什么
    # proficiency只是记录上次复习的结果，不影响是否需要复习
    review_items_list = []
    
    for item in review_items_query:
        # 计算逾期天数
        overdue_days = 0
        if item.next_review_date and item.next_review_date < reviewDate:
            overdue_days = (reviewDate - item.next_review_date).days
        
        # 判断复习类型
        is_regular = item.next_review_date and item.next_review_date <= reviewDate
        is_extra = item.needs_extra_review and item.extra_review_since and item.extra_review_since <= reviewDate
        
        # 统计当前周期的不熟悉次数（兼容记录中存的是索引或天数）
        unfamiliar_count = 0
        if item.unfamiliar_history:
            try:
                intervals = [0, 1, 2, 4, 7, 15, 30, 90, 180]
                # 如果 current_interval 是天数，尝试获取对应的索引；否则如果是索引，则获取对应天数
                if item.current_interval in intervals:
                    cur_day = item.current_interval
                    try:
                        cur_idx = intervals.index(cur_day)
                    except ValueError:
                        cur_idx = None
                else:
                    # 可能是索引值
                    try:
                        cur_idx = int(item.current_interval)
                        cur_day = intervals[cur_idx] if 0 <= cur_idx < len(intervals) else item.current_interval
                    except Exception:
                        cur_idx = None
                        cur_day = item.current_interval

                unfamiliar_count = sum(
                    1 for record in item.unfamiliar_history
                    if record.get('interval') == cur_day or record.get('interval') == cur_idx
                )
            except Exception:
                unfamiliar_count = 0

        # 如果该 item 在今天已经被点评（unfamiliar_history 里有今天的记录），则标记为已点评但仍显示
        reviewed_today = False
        last_review_type = None
        if item.unfamiliar_history:
            for record in item.unfamiliar_history:
                try:
                    if record.get('date') == reviewDate.isoformat():
                        reviewed_today = True
                        last_review_type = record.get('review_type')
                        break
                except Exception:
                    continue
        
        # 计算如果点YES，下次复习日期是什么
        next_review_after_yes = None
        if is_regular:
            # 正式复习：会进入下一个间隔
            intervals = [0, 1, 2, 4, 7, 15, 30, 90, 180]
            try:
                current_index = intervals.index(item.current_interval)
                if current_index < len(intervals) - 1:
                    next_interval = intervals[current_index + 1]
                    current_interval_value = intervals[current_index]
                    days_until_next = next_interval - current_interval_value
                    next_review_after_yes = reviewDate + timedelta(days=days_until_next)
                else:
                    # 已经是最后一个间隔
                    next_review_after_yes = reviewDate + timedelta(days=365)
            except ValueError:
                next_review_after_yes = reviewDate + timedelta(days=1)
        else:
            # 只是额外复习：next_review_date不变
            next_review_after_yes = item.next_review_date
        
        # 生成详细页面URL
        detail_url = reverse('item-detail', args=[item.pk])
        
        review_items_list.append({
            'item': item,
            'category_name': item.category.name,
            'interval_day': item.current_interval,
            'next_review_date': item.next_review_date or reviewDate,
            'next_review_after_yes': next_review_after_yes,
            'overdue_days': overdue_days,
            'is_regular': is_regular,
            'is_extra': is_extra,
            'unfamiliar_count': unfamiliar_count,
            'reviewed_today': reviewed_today,
            'last_review_type': last_review_type,
            'detail_url': detail_url
        })
    
    # 排序：额外复习优先，然后按类别和间隔
    review_items_list.sort(key=lambda x: (not x['is_extra'], x['category_name'], x['interval_day']))
    
    # 分页处理
    paginator = Paginator(review_items_list, per_page)
    page_obj = paginator.get_page(page_number)
    
    # 构建context
    context = {
        'page_obj': page_obj,
        'reviewdate': reviewDate,
        'show_mastered': show_mastered,
        'per_page': per_page
    }
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return HttpResponse(render_to_string('review_day.html', context, request))

    return render(request, 'review_day.html', context)


@login_required
def ReviewFeedbackYes(request):
    """
    更新指定 Item 的 proficiency 为 MASTERED（熟练）。
    根据新逻辑：正式复习进入下一周期，额外复习只清除标记。
    """
    try:
        if request.method == "POST":
            # 解析请求数据
            data = json.loads(request.body.decode("utf-8"))
            item_id = data.get('id')
            review_date_str = data.get('date')
            
            if review_date_str:
                review_date = datetime.strptime(str(review_date_str), '%Y-%m-%d').date()
            else:
                review_date = datetime.today().date()

            # 获取当前用户的 Item
            curword = Item.objects.get(user=request.user, id=item_id)
            
            # 判断是否为正式复习
            is_regular_review = curword.next_review_date and curword.next_review_date <= review_date
            
            # 更新 proficiency 为 MASTERED
            curword.proficiency = Proficiency.MASTERED
            
            if is_regular_review:
                # 正式复习：进入下一个周期
                intervals = [0, 1, 2, 4, 7, 15, 30, 90, 180]
                try:
                    current_index = intervals.index(curword.current_interval)
                    if current_index < len(intervals) - 1:
                        # 进入下一个间隔
                        next_index = current_index + 1
                        next_interval = intervals[next_index]
                        current_interval_value = intervals[current_index]
                        
                        # 更新间隔
                        curword.current_interval = next_interval
                        
                        # 计算距离下次复习的天数 = 下个间隔 - 当前间隔
                        # 例如：Day 0→Day 1: 1-0=1天后
                        #      Day 1→Day 2: 2-1=1天后
                        #      Day 2→Day 4: 4-2=2天后
                        days_until_next = next_interval - current_interval_value
                        curword.next_review_date = review_date + timedelta(days=days_until_next)
                    else:
                        # 已经是最后一个间隔（Day 365），保持365天周期
                        curword.next_review_date = review_date + timedelta(days=365)
                except ValueError:
                    # 如果current_interval不在列表中，按保守策略重置为Day 1，并且不尝试删除历史记录
                    curword.current_interval = 1
                    curword.next_review_date = review_date + timedelta(days=1)

                # 清除额外复习标记
                curword.needs_extra_review = False
                curword.extra_review_since = None

                # 仅在成功解析出本次周期标识时，才从 unfamiliar_history 中移除对应记录；否则保留原记录
                try:
                    uh = curword.unfamiliar_history or []
                    # 仅在解析到本次周期标识时进行有选择地移除；使用类型安全的比较
                    if 'current_interval_value' in locals() or 'current_index' in locals():
                        civ = locals().get('current_interval_value')
                        cidx = locals().get('current_index')

                        def matches_cycle(rec_val, civ_val, cidx_val):
                            try:
                                # 尝试以整数比较
                                rv = int(rec_val)
                            except Exception:
                                rv = rec_val
                            return rv == civ_val or rv == cidx_val

                        filtered = [r for r in uh if not matches_cycle(r.get('interval'), civ, cidx)]
                        curword.unfamiliar_history = filtered
                    else:
                        # 未能解析周期标识，保留原始历史以避免误删
                        curword.unfamiliar_history = uh
                except Exception:
                    # 出现任何意外时，保留原始历史（不要盲目清空）
                    pass
            else:
                # 只是额外复习：清除额外复习标记，正式日期不变
                pass
            
            # 无论是正式复习还是额外复习，点YES后都清除额外复习标记
            curword.needs_extra_review = False
            curword.extra_review_since = None
            
            # 幂等性：如果当天已存在相同日期的记录，返回已点评信息而不重复修改
            try:
                if curword.unfamiliar_history:
                    for r in curword.unfamiliar_history:
                        if r.get('date') == review_date.isoformat():
                            curword.save()
                            return JsonResponse({
                                'success': True,
                                'message': 'Already reviewed today',
                                'mastery': curword.get_proficiency_display(),
                                'reviewed_today': True,
                            })
            except Exception:
                pass

            curword.save()

            # ========== 积分系统集成：复习奖励 ==========
            try:
                from django.utils import timezone
                from EAW.models import UserPoints, PointHistory, UserPointsConfig, UserStreak

                points_account, created = UserPoints.objects.get_or_create(
                    user=request.user,
                    defaults={'current_points': 0}
                )

                # 检查今天是否已经对该单词获得过积分
                today = timezone.now().date()
                already_earned_today = PointHistory.objects.filter(
                    user=request.user,
                    reference_id=f"review_{curword.id}",
                    created_at__date=today
                ).exists()

                if not already_earned_today:
                    # 基础积分：复习完成
                    points_earned = 1
                    points_account.add_points(
                        points=points_earned,
                        reason=f"完成单词复习：{curword.item}",
                        reference_id=f"review_{curword.id}"
                    )

                    # 更新连续学习记录
                    streak, _ = UserStreak.objects.get_or_create(user=request.user)
                    streak.update_study_streak(today)

                    # 检查连续学习奖励
                    config, _ = UserPointsConfig.objects.get_or_create(user=request.user)
                    streak_reward = streak.check_streak_reward(config)
                    if streak_reward:
                        points_account.add_points(
                            points=streak_reward,
                            reason=f"连续学习{config.streak_reward_days}天奖励",
                            reference_id=f"streak_{streak.current_streak}"
                        )
                        points_earned += streak_reward

                    # 返回包含积分信息的响应
                    return JsonResponse({
                        'success': True,
                        'message': 'Proficiency updated to MASTERED.',
                        'mastery': curword.get_proficiency_display(),
                        'reviewed_today': True,
                        'points_earned': points_earned,
                        'current_points': points_account.current_points,
                        'current_streak': streak.current_streak
                    })
            except Exception as points_error:
                # 积分系统失败不应影响主流程，记录日志但不抛出异常
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Points system error: {str(points_error)}")

            # 返回正常响应（没有积分奖励或积分系统出错）
            return JsonResponse({
                'success': True,
                'message': 'Proficiency updated to MASTERED.',
                'mastery': curword.get_proficiency_display(),
                'reviewed_today': True,
            })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def ReviewFeedbackNo(request):
    """
    更新指定 Item 的 proficiency 为 UNFAMILIAR（不熟练）。
    根据新逻辑：记录不熟悉，如果达到3次则重启周期，否则进入下一周期并设置额外复习。
    """
    try:
        if request.method == "POST":
            # 解析请求数据
            data = json.loads(request.body.decode("utf-8"))
            item_id = data.get('id')
            review_date_str = data.get('date')
            
            if review_date_str:
                review_date = datetime.strptime(str(review_date_str), '%Y-%m-%d').date()
            else:
                review_date = datetime.today().date()

            # 获取当前用户的 Item
            curword = Item.objects.get(user=request.user, id=item_id)
            
            # 判断是否为正式复习
            is_regular_review = curword.next_review_date and curword.next_review_date <= review_date
            
            # 记录不熟悉
            if not curword.unfamiliar_history:
                curword.unfamiliar_history = []
            
            # 规范化存储：尽量把 interval 存为间隔天数（比如 0,1,2,4...），兼容旧的索引值
            try:
                intervals = [0, 1, 2, 4, 7, 15, 30, 90, 180]
                if curword.current_interval in intervals:
                    interval_value = curword.current_interval
                else:
                    # 如果 current_interval 看起来像索引，尝试转换为对应天数
                    try:
                        idx = int(curword.current_interval)
                        interval_value = intervals[idx] if 0 <= idx < len(intervals) else curword.current_interval
                    except Exception:
                        interval_value = curword.current_interval
            except Exception:
                interval_value = curword.current_interval

            # 在添加前检查是否已存在当天的记录以保证幂等性
            exists_today = False
            try:
                if curword.unfamiliar_history:
                    for r in curword.unfamiliar_history:
                        if r.get('date') == review_date.isoformat():
                            exists_today = True
                            break
            except Exception:
                exists_today = False

            if not exists_today:
                curword.unfamiliar_history.append({
                    "date": review_date.isoformat(),
                    "interval": interval_value,
                    "review_type": "regular" if is_regular_review else "extra"
                })
            
            # 更新 proficiency 为 UNFAMILIAR
            curword.proficiency = Proficiency.UNFAMILIAR
            
            # 统计当前周期的不熟悉次数（基于规范化的 interval_value）
            try:
                current_cycle_count = sum(
                    1 for record in curword.unfamiliar_history
                    if record.get('interval') == interval_value
                )
            except Exception:
                current_cycle_count = 0
            
            message = 'Proficiency updated to UNFAMILIAR.'
            
            if current_cycle_count >= 3:
                # 达到3次，重新开始周期
                old_interval = curword.current_interval
                curword.current_interval = 0
                curword.next_review_date = review_date  # Day 0就是今天
                curword.unfamiliar_history = []
                curword.needs_extra_review = False
                curword.extra_review_since = None
                message = f'该单词在Day {old_interval}已标记3次不熟悉，将从今天重新开始复习周期（Day 0）。'
                
            elif is_regular_review:
                # 正式复习日点NO：不进入下一周期，保持当前周期 + 设置额外复习
                # 等到点YES后才进入下一周期
                # 设置额外复习
                curword.needs_extra_review = True
                curword.extra_review_since = review_date + timedelta(days=1)
                # next_review_date 保持不变，继续在当前周期
            
            # else: 额外复习期间点NO，继续额外复习，next_review_date 不变
            
            curword.save()

            # 如果当天已存在记录，向前端说明为已点评
            resp = {
                'success': True,
                'message': message,
                'mastery': curword.get_proficiency_display(),
                'unfamiliar_count': current_cycle_count,
                'reviewed_today': exists_today or True,
            }

            return JsonResponse(resp)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def ReviewFeedbackReset(request):
    """
    重置指定 Item 的复习周期，从Day 0重新开始。
    """
    try:
        if request.method == "POST":
            # 解析请求数据
            data = json.loads(request.body.decode("utf-8"))
            item_id = data.get('id')

            # 获取当前用户的 Item
            curword = Item.objects.get(user=request.user, id=item_id)

            # 重置复习周期
            today = datetime.today().date()
            curword.current_interval = 0
            curword.next_review_date = today  # 今天就开始Day 0
            curword.proficiency = Proficiency.UNFAMILIAR
            curword.unfamiliar_history = []
            curword.needs_extra_review = False
            curword.extra_review_since = None
            curword.save()

            return JsonResponse({
                'success': True,
                'message': 'Review cycle has been reset. Starting from Day 0.',
                'mastery': curword.get_proficiency_display()
            })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

#BAIDU API调试页面
def translate_test(request):
    return render(request, 'translate.html')  # 渲染index.html页面
@csrf_exempt
def translate(request):
    if request.method == 'POST':
        try:
            # 获取前端传来的查询词
            data = json.loads(request.body)
            query = data.get('query', '')

            if not query:
                return JsonResponse({'success': False, 'message': 'No query provided'})

            # 调用百度翻译接口
            result = baidu_translate(query)
            
            if result:
                return JsonResponse({'success': True, 'result': result})
            else:
                return JsonResponse({'success': False, 'message': 'Translation failed'})
        
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})
def check_api_keys_view(request):
    """
    检查是否配置了百度 API 密钥
    :return: JsonResponse
    """
    if check_api_keys():
        return JsonResponse({"success": True, "message": "百度 API 密钥已配置"})
    else:
        return JsonResponse({"success": False, "message": "百度 API 密钥未配置"}, status=400)
    

@login_required
def InputView(request):
    if request.method == 'POST':
        form = InputForm(request.POST, user=request.user)  # 传递当前用户
        if form.is_valid():
            data = {
                'input_date': form.cleaned_data['input_date'],
                'category': form.cleaned_data['category'].name,
                'input': form.cleaned_data['input']
            }
            split = data['input'].split('\r\n')
            category_object = Category.objects.get(name=data['category'], user=request.user)  # 仅查找当前用户的类别

            # 获取是否勾选了翻译复选框，并且类别为"单词"
            translate = 'translate' in request.POST and data['category'] == '单词'

            for item in split:
                explain_txt = ''
                result_dict = None
                translated_content = ''
                simple_meaning = ''
                item_name = item
                item_name, explain_txt = split_string(item)  # 如果有拆分功能
                item_name = item_name.strip()
                # 初始化 phonetic_am 和 phonetic_en 为 None
                phonetic_am = phonetic_en = None
                src_tts = None

                # 如果勾选了 "获取释义" 复选框，则调用百度翻译函数
                if translate:
                    result_dict = baidu_translate(item_name)
                    if result_dict:  # 如果返回的字典非空
                        # 从 result_dict 中提取各个部分
                        phonetic = result_dict.get('phonetic', [])
                        phonetic_am = phonetic[1] if len(phonetic) > 1 else None  # 美式音标
                        phonetic_en = phonetic[0] if len(phonetic) > 0 else None  # 英式音标
                        src_tts = result_dict.get('src_tts', None)  # TTS URL
                        translated_content = result_dict.get('parts_and_means', [])  # 词性和释义
                        simple_meaning = result_dict.get('simple_meaning', [])  # 简明释义

                        # 拼接解释文本
                    # 确保是字符串并避免空行
                        if translated_content:
                            if explain_txt:  # 如果原来已有内容，才添加换行
                                explain_txt += "\n\n"
                            explain_txt += "\n".join([str(item) for item in translated_content])  # 拼接详细释义

                        # 如果有 translated_content 或音标，则不存储 simple_meaning
                        if not translated_content:
                            # 如果没有翻译内容才拼接简明释义
                            if simple_meaning:
                                if explain_txt:  # 如果原来已有内容，才添加换行
                                    explain_txt += "\n\n"
                                explain_txt += "\n".join([str(item) for item in simple_meaning])  # 拼接简明释义
                    else:
                        phonetic_am = phonetic_en = src_tts = None

                # 创建 Item 实例，并保存到数据库
                Item.objects.create(
                    user=request.user,
                    item=item_name,
                    inputDate=data['input_date'],
                    initDate=data['input_date'],
                    category=category_object,
                    content=explain_txt,
                    src_tts=src_tts if translate else None,  # 如果未勾选翻译，TTS 地址为 None
                    us_phonetic=phonetic_am,  # 存储美式音标
                    uk_phonetic=phonetic_en,   # 存储英式音标
                    # 新增复习系统字段
                    current_interval=0,
                    next_review_date=data['input_date'],  # Day 0从录入日期开始
                    needs_extra_review=False,
                    unfamiliar_history=[]
                )

            return redirect(reverse('item-list'))  # 重定向到项列表页面
    else:
        form = InputForm(user=request.user)  # 传递当前用户

    return render(request, 'input.html', {'form': form})



def split_string(s):
    # 查找第一个出现的英文冒号或中文冒号的位置
    pos = s.find(":")
    pos_cn = s.find("：")
    
    # 找到最先出现的冒号位置
    if pos == -1 or (pos_cn != -1 and pos_cn < pos):
        pos = pos_cn
    
    # 如果没有冒号，返回原字符串和空字符串
    if pos == -1:
        return s, ""
    
    # 根据位置分割字符串
    return s[:pos], s[pos + 1:]


@login_required
def export_user_data_to_excel(request):
    # 获取当前用户数据
    user = request.user
    items = Item.objects.filter(user=user)

    # 创建一个新的 Excel 工作簿
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "User Data"

    # 写入表头
    headers = ["Item", "Content", "Input Date", "Init Date", "Proficiency", "Category", "TTS URL", "US Phonetic", "UK Phonetic"]
    sheet.append(headers)

    # 写入用户数据
    for item in items:
        sheet.append([
            item.item,
            item.content,
            item.inputDate,
            item.initDate,
            item.get_proficiency_display(),
            item.category.name if item.category else "",
            item.src_tts,
            item.us_phonetic,
            item.uk_phonetic,
        ])

    # 创建 HTTP 响应
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="user_data.xlsx"'

    # 将工作簿保存到响应中
    workbook.save(response)

    return response
    


@login_required
def import_items_from_excel(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]
        user = request.user

        # 检查是否选择了“获取释义”选项
        fetch_definitions = "fetch_definitions" in request.POST

        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
        except Exception as e:
            messages.error(request, f"文件读取失败: {str(e)}")
            return render(request, "import_data.html")

        headers = [cell.value for cell in sheet[1]]
        required_columns = ["Item"]
        if not all(col in headers for col in required_columns):
            messages.error(request, "文件格式错误，缺少必要的列。")
            return render(request, "import_data.html")

        column_index = {header: headers.index(header) for header in headers}
        items_to_create = []
        errors = []

        # Proficiency 字段的映射
        proficiency_map = {
            "Unfamiliar": Proficiency.UNFAMILIAR,
            "Mastered": Proficiency.MASTERED,
        }

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                item_name = row[column_index["Item"]]
                if not item_name:
                    errors.append(f"第 {row_idx} 行缺少 Item 字段，已跳过。")
                    continue

                content_index = column_index.get("Content")
                content = row[content_index] if content_index is not None else ""
                # 替换 _x000D_ 字符为换行符，先检查是否为 None
                if content:
                    content = content.replace("_x000D_", "\n")
                else:
                    content = ""  # 如果 content 为 None，赋空字符串

                input_date_index = column_index.get("Input Date")
                input_date = row[input_date_index] if input_date_index is not None else now().date()

                init_date_index = column_index.get("Init Date")
                init_date = row[init_date_index] if init_date_index is not None else now().date()

                # 处理 Proficiency 字段
                proficiency_name = row[column_index.get("Proficiency", None)] or "Unfamiliar"
                proficiency_degree = proficiency_map.get(proficiency_name, Proficiency.UNFAMILIAR)

                category_index = column_index.get("Category")
                category_name = row[category_index] if category_index is not None else ""

                # 获取分类对象
                category = None
                if category_name:
                    categories = Category.objects.filter(name=category_name, user=user)
                    if categories.exists():
                        category = categories.first()
                    else:
                        category = Category.objects.create(name=category_name, user=user)

                # 初始化这些字段为默认值
                src_tts = ""
                phonetic_am = ""
                phonetic_en = ""

                # 只有类别为“单词”的条目才调用获取翻译的功能
                if fetch_definitions and category and category.name == "单词":
                    updated_content, src_tts, phonetic_am, phonetic_en = fetch_and_merge_translation(item_name, content)
                    # 更新内容
                    content = updated_content

                item = Item(
                    user=user,
                    item=item_name,
                    content=content,
                    inputDate=input_date,
                    initDate=init_date,
                    proficiency=proficiency_degree,
                    category=category,
                    src_tts=src_tts,
                    us_phonetic=phonetic_am,
                    uk_phonetic=phonetic_en,
                )
                items_to_create.append(item)

            except Exception as e:
                errors.append(f"第 {row_idx} 行处理失败: {str(e)}")

        try:
            with transaction.atomic():
                Item.objects.bulk_create(items_to_create)
            success_message = f"导入完成。成功导入 {len(items_to_create)} 条记录，{len(errors)} 条记录跳过。"
            messages.success(request, success_message)
        except Exception as e:
            messages.error(request, f"保存失败: {str(e)}")

        return render(request, "import_data.html", {
            "import_results": {
                "success_count": len(items_to_create),
                "errors": errors,
            }
        })

    messages.error(request, "请求无效，请上传文件。")
    return render(request, "import_data.html")




@staticmethod
def compare_lines(existing_lines, new_lines):
    existing = [line.strip() for line in existing_lines.splitlines() if line.strip()]
    new = [line.strip() for line in new_lines.splitlines() if line.strip()]

    result = []
    existing_set = set(existing)

    # 逐行比较
    for line in existing:
        result.append(line)

    for new_line in new:
        if not any(difflib.SequenceMatcher(None, new_line, e_line).ratio() > 0.8 for e_line in existing_set):
            result.append(new_line)

    return "\n".join(result)
def about(request):
    return render(request, 'about.html')  # 渲染 about.html 页面

def readme_view(request):
    # 使用 BASE_DIR 获取 README.md 文件的路径
    readme_path = os.path.join(settings.BASE_DIR, 'README.md')

    # 读取文件内容
    with open(readme_path, 'r', encoding='utf-8') as f:
        readme_content = f.read()

    # 将 Markdown 转换为 HTML，并启用 fenced_code 扩展
    html_content = markdown.markdown(readme_content, extensions=['fenced_code'])

    # 渲染模板
    return render(request, 'readme.html', {'content': html_content})


def dev_log(request):
    """
    渲染仓库根下的 DEVELOPMENT_LOG.md 为 HTML 并显示（与 readme_view 风格一致）。
    """
    devlog_path = os.path.join(settings.BASE_DIR, 'DEVELOPMENT_LOG.md')
    if os.path.exists(devlog_path):
        with open(devlog_path, 'r', encoding='utf-8') as f:
            md = f.read()
        html_content = markdown.markdown(md, extensions=['fenced_code', 'tables'])
    else:
        html_content = '<p>No development log found.</p>'

    return render(request, 'dev_log.html', {'content': html_content})


# ==================== DeepSeek API 相关视图 ====================

@login_required
def deepseek_config_view(request):
    """DeepSeek API 配置页面"""
    # 获取或创建用户的配置
    config, created = DeepSeekConfig.objects.get_or_create(
        user=request.user,
        defaults={
            'model': 'deepseek-chat',
            'temperature': 1.0,
            'system_prompt': '你是英语词典助手。输入：英文单词。输出：JSON格式包含uk_phonetic(英音标)、us_phonetic(美音标)、meaning(中文释义)、example_sentences(2-3个例句数组，每个包含english和chinese字段)。目标用户：小学5年级。严格按照JSON格式输出，不要添加任何其他文字或markdown代码块标记。',
            'is_active': True
        }
    )
    
    if request.method == 'POST':
        form = DeepSeekConfigForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, 'DeepSeek 配置已成功保存！')
            return redirect('deepseek-config')
        else:
            messages.error(request, '配置保存失败，请检查输入。')
    else:
        form = DeepSeekConfigForm(instance=config)
    
    return render(request, 'deepseek_config.html', {'form': form})


@login_required
def deepseek_query_view(request):
    """DeepSeek 查询页面"""
    if request.method == 'GET':
        return render(request, 'deepseek_query.html')
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            word = data.get('word', '').strip()
            
            if not word:
                return JsonResponse({'success': False, 'error': '请输入单词'})
            
            # 调用 DeepSeek API
            result = call_deepseek_api(word, user=request.user)
            
            if not result:
                return JsonResponse({'success': False, 'error': 'API 调用失败或返回空结果'})
            
            # 解析结果
            uk_phonetic = result.get('phonetic', [])[0] if result.get('phonetic') else ''
            us_phonetic = result.get('phonetic', [])[1] if len(result.get('phonetic', [])) > 1 else ''
            
            # 提取简明释义
            simple_meaning = result.get('simple_meaning', [])
            meaning = simple_meaning[0].replace('简明释义: ', '') if simple_meaning else ''
            
            # 提取例句（从 parts_and_means）
            parts_and_means = result.get('parts_and_means', [])
            example_sentences = []
            
            if parts_and_means:
                # 解析例句
                content = '\n'.join(parts_and_means)
                lines = content.split('\n')
                current_example = {}
                
                for line in lines:
                    line = line.strip()
                    if line.startswith('例句'):
                        if current_example:
                            example_sentences.append(current_example)
                        current_example = {'english': line.split(': ', 1)[1] if ': ' in line else line}
                    elif line.startswith('翻译'):
                        current_example['chinese'] = line.split(': ', 1)[1] if ': ' in line else line
                
                if current_example and 'chinese' in current_example:
                    example_sentences.append(current_example)
            
            response_data = {
                'success': True,
                'data': {
                    'uk_phonetic': uk_phonetic,
                    'us_phonetic': us_phonetic,
                    'meaning': meaning,
                    'example_sentences': example_sentences
                }
            }
            
            return JsonResponse(response_data)
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'JSON 解析失败'})
        except Exception as e:
            logger.error(f"DeepSeek query error: {e}")
            return JsonResponse({'success': False, 'error': str(e)})


@login_required
def deepseek_save_view(request):
    """保存 DeepSeek 查询结果到数据库"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            word = data.get('word', '').strip()
            query_data = data.get('data', {})
            category_id = data.get('category_id')

            if not word:
                return JsonResponse({'success': False, 'error': '单词不能为空'})

            # 获取分类：优先使用用户指定的分类
            category = None
            if category_id:
                try:
                    category = Category.objects.get(id=category_id, user=request.user)
                except Category.DoesNotExist:
                    return JsonResponse({'success': False, 'error': '指定的分类不存在'})

            # 如果没有指定分类，获取默认分类"单词"（向后兼容）
            if not category:
                category = Category.objects.filter(
                    user=request.user,
                    name='单词'
                ).first()

                # 如果没有找到"单词"分类，创建或使用第一个分类
                if not category:
                    category, _ = Category.objects.get_or_create(
                        user=request.user,
                        name='单词',
                        defaults={'is_default': True, 'sort_order': 0}
                    )
            
            # 构建内容
            content_parts = []
            
            # 添加释义
            if query_data.get('meaning'):
                content_parts.append(f"释义: {query_data['meaning']}")
            
            # 添加例句
            if query_data.get('example_sentences'):
                content_parts.append("\n例句:")
                for idx, example in enumerate(query_data['example_sentences'], 1):
                    content_parts.append(f"{idx}. {example.get('english', '')}")
                    content_parts.append(f"   {example.get('chinese', '')}")
            
            content = '\n'.join(content_parts)
            
            # 获取今天的日期
            today = now().date()
            
            # 创建 Item
            item = Item.objects.create(
                user=request.user,
                item=word,
                content=content,
                inputDate=today,
                initDate=today,
                proficiency=Proficiency.UNFAMILIAR,
                category=category,
                uk_phonetic=query_data.get('uk_phonetic', ''),
                us_phonetic=query_data.get('us_phonetic', ''),
                src_tts=f"https://translate.google.com/translate_tts?ie=UTF-8&client=tw-ob&tl=en&q={word}",
                # 新增复习系统字段
                current_interval=0,
                next_review_date=today,  # Day 0从今天开始
                needs_extra_review=False,
                unfamiliar_history=[]
            )
            
            return JsonResponse({
                'success': True,
                'message': '保存成功',
                'item_id': item.id
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'JSON 解析失败'})
        except Exception as e:
            logger.error(f"DeepSeek save error: {e}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': '仅支持 POST 请求'})


@login_required
def get_last_item_category_view(request):
    """获取用户最后保存的词汇的分类（用于设置默认分类）"""
    if request.method == 'GET':
        try:
            # 获取用户最后保存的词汇
            last_item = Item.objects.filter(user=request.user).order_by('-inputDate').first()

            if last_item and last_item.category:
                return JsonResponse({
                    'success': True,
                    'category_id': last_item.category.id,
                    'category_name': last_item.category.name
                })
            else:
                return JsonResponse({'success': False, 'message': 'No items found'})

        except Exception as e:
            logger.error(f"Error getting last item category: {e}")
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': '仅支持 GET 请求'})


def check_deepseek_keys_view(request):
    """检查 DeepSeek API Key 是否配置"""
    configured = check_deepseek_keys()
    return JsonResponse({'configured': configured})


# ==================== 积分系统视图 ====================

@login_required
def points_market_view(request):
    """积分商城页面"""
    from EAW.models import UserPoints, UserPointsConfig, UserStreak, PointRedemption, PointHistory

    user = request.user

    # 获取用户积分和配置
    points_account, _ = UserPoints.objects.get_or_create(user=user)
    config, _ = UserPointsConfig.objects.get_or_create(user=user)
    streak, _ = UserStreak.objects.get_or_create(user=user)

    # 检查今天是否已签到
    today = now().date()
    has_checkin_today = PointHistory.objects.filter(
        user=user,
        reason__startswith='每日签到',
        created_at__date=today
    ).exists()

    # 最近兑换记录
    recent_redemptions = PointRedemption.objects.filter(
        user=user
    ).order_by('-created_at')[:10]

    context = {
        'points_account': points_account,
        'config': config,
        'streak': streak,
        'has_checkin_today': has_checkin_today,
        'recent_redemptions': recent_redemptions
    }
    return render(request, 'points_market.html', context)


@login_required
def points_config_view(request):
    """用户积分配置页面"""
    from EAW.models import UserPoints, UserPointsConfig, UserStreak

    config, _ = UserPointsConfig.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        # 更新配置
        config.minutes_per_point = float(request.POST.get('minutes_per_point', 1.0))
        config.redemption_step = int(request.POST.get('redemption_step', 10))
        config.min_redemption_minutes = int(request.POST.get('min_redemption_minutes', 1))
        config.max_redemption_minutes = int(request.POST.get('max_redemption_minutes', 300))
        config.daily_checkin_enabled = request.POST.get('daily_checkin') == 'on'
        config.daily_checkin_points = int(request.POST.get('daily_checkin_points', 1))
        config.streak_reward_enabled = request.POST.get('streak_reward') == 'on'
        config.streak_reward_points = int(request.POST.get('streak_reward_points', 5))
        config.streak_reward_days = int(request.POST.get('streak_reward_days', 7))
        config.save()

        messages.success(request, '配置已更新')
        return redirect('points-config')

    points_account, _ = UserPoints.objects.get_or_create(user=request.user)
    streak, _ = UserStreak.objects.get_or_create(user=request.user)

    context = {
        'config': config,
        'points_account': points_account,
        'streak': streak
    }
    return render(request, 'points_config.html', context)


@login_required
@require_http_methods(["POST"])
def redeem_game_time_view(request):
    """兑换游戏时长API"""
    from EAW.models import UserPoints, UserPointsConfig, PointRedemption

    try:
        data = json.loads(request.body)
        minutes = int(data.get('minutes', 0))

        if minutes <= 0:
            return JsonResponse({'success': False, 'message': '请选择有效的时长'})

        user = request.user

        # 获取用户配置
        config, _ = UserPointsConfig.objects.get_or_create(user=user)

        # 验证兑换时长范围
        if minutes < config.min_redemption_minutes:
            return JsonResponse({
                'success': False,
                'message': f'最少兑换{config.min_redemption_minutes}分钟'
            })
        if minutes > config.max_redemption_minutes:
            return JsonResponse({
                'success': False,
                'message': f'最多兑换{config.max_redemption_minutes}分钟'
            })

        # 计算所需积分
        points_needed = int(minutes / config.minutes_per_point)

        # 获取积分账户
        points_account, _ = UserPoints.objects.get_or_create(user=user)

        # 检查积分是否足够
        if points_account.current_points < points_needed:
            return JsonResponse({
                'success': False,
                'message': f'积分不足！需要{points_needed}积分，当前{points_account.current_points}积分'
            })

        # 使用事务扣除积分并创建兑换记录
        with transaction.atomic():
            # 扣除积分
            points_account.spend_points(
                points=points_needed,
                reason=f"兑换{minutes}分钟游戏时长",
                reference_id=f"redeem_{now().timestamp()}"
            )

            # 创建兑换记录
            PointRedemption.objects.create(
                user=user,
                points_spent=points_needed,
                game_minutes=minutes,
                exchange_rate=config.minutes_per_point,
                status='COMPLETED'
            )

        return JsonResponse({
            'success': True,
            'message': f'成功兑换{minutes}分钟游戏时长！消费{points_needed}积分',
            'remaining_points': points_account.current_points
        })

    except Exception as e:
        logger.error(f"Redemption error: {str(e)}")
        return JsonResponse({'success': False, 'message': f'兑换失败：{str(e)}'})


@login_required
@require_http_methods(["POST"])
def daily_checkin_view(request):
    """每日签到API"""
    from EAW.models import UserPoints, UserPointsConfig, UserStreak, PointHistory

    try:
        user = request.user
        today = now().date()

        # 检查今天是否已签到
        existing = PointHistory.objects.filter(
            user=user,
            reason__startswith='每日签到',
            created_at__date=today
        ).exists()

        if existing:
            return JsonResponse({'success': False, 'message': '今天已经签到过了'})

        # 获取配置
        config, _ = UserPointsConfig.objects.get_or_create(user=user)

        if not config.daily_checkin_enabled:
            return JsonResponse({'success': False, 'message': '签到功能未启用'})

        # 更新连续签到
        streak, _ = UserStreak.objects.get_or_create(user=user)
        streak.update_checkin_streak(today)

        # 给予积分奖励
        points_account, _ = UserPoints.objects.get_or_create(user=user)
        points_account.add_points(
            points=config.daily_checkin_points,
            reason=f"每日签到（连续{streak.current_checkin_streak}天）",
            reference_id=f"checkin_{today}"
        )

        return JsonResponse({
            'success': True,
            'message': f'签到成功！获得{config.daily_checkin_points}积分',
            'checkin_streak': streak.current_checkin_streak,
            'points_earned': config.daily_checkin_points,
            'current_points': points_account.current_points
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'签到失败：{str(e)}'})


@login_required
def gobang_game(request):
    """
    五子棋游戏页面
    点击开始游戏时扣除5积分
    """
    user = request.user
    from EAW.models import UserPoints
    points_account, _ = UserPoints.objects.get_or_create(user=user)

    # 添加调试日志
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f'[GOBANG PAGE] User: {user.username}, Current: {points_account.current_points}, '
                  f'Total Spent: {points_account.total_spent}, '
                  f'Total Earned: {points_account.total_earned}')

    return render(request, 'gobang.html', {
        'remaining_points': points_account.current_points,
        'total_spent': points_account.total_spent,  # 累计消费积分
        'points_per_game': 5,  # 每次游戏消耗5积分
        'debug_info': f'DB total_spent={points_account.total_spent}, current_points={points_account.current_points}'
    })


@login_required
@require_http_methods(["POST"])
def gobang_start_game_api(request):
    """
    五子棋开始游戏API - 扣除5积分
    """
    from EAW.models import UserPoints

    try:
        user = request.user
        REQUIRED_POINTS = 5

        # 获取积分账户
        points_account, _ = UserPoints.objects.get_or_create(user=user)

        # 检查积分是否足够
        if points_account.current_points < REQUIRED_POINTS:
            return JsonResponse({
                'success': False,
                'message': f'积分不足！开始游戏需要{REQUIRED_POINTS}积分，当前{points_account.current_points}积分'
            })

        # 保存扣除前的累计消费
        old_total_spent = points_account.total_spent

        # 使用事务扣除积分
        with transaction.atomic():
            points_account.spend_points(
                points=REQUIRED_POINTS,
                reason="五子棋游戏",
                reference_id=f"gobang_{now().timestamp()}"
            )

        # 重新获取更新后的数据
        points_account.refresh_from_db()

        return JsonResponse({
            'success': True,
            'message': f'开始游戏成功！扣除{REQUIRED_POINTS}积分',
            'remaining_points': points_account.current_points,
            'total_spent': points_account.total_spent,  # 更新后的累计消费
            'points_per_game': REQUIRED_POINTS  # 每局消耗
        })

    except Exception as e:
        logger.error(f"Gobang start game error: {str(e)}")
        return JsonResponse({'success': False, 'message': f'操作失败：{str(e)}'})

    # ===== 以下积分扣除逻辑在测试期间已注释 =====
    # # 配置每次游戏需要的积分（默认 10 积分）
    # REQUIRED_POINTS = 10
    #
    # # 检查积分是否足够
    # if points_account.current_points < REQUIRED_POINTS:
    #     return render(request, 'gobang_insufficient_points.html', {
    #         'required_points': REQUIRED_POINTS,
    #         'current_points': points_account.current_points
    #     })
    #
    # # 使用事务扣除积分
    # with transaction.atomic():
    #     points_account.spend_points(
    #         points=REQUIRED_POINTS,
    #         reason="五子棋游戏",
    #         reference_id=f"gobang_{request.session.session_key}"
    #     )
    #
    # # 渲染游戏页面
    # return render(request, 'gobang.html', {
    #     'points_spent': REQUIRED_POINTS,
    #     'remaining_points': points_account.current_points
    # })


@login_required
def points_history_view(request):
    """积分历史记录页面"""
    from EAW.models import PointHistory

    user = request.user
    page_number = request.GET.get('page', 1)
    per_page = int(request.GET.get('per_page', 20))
    change_type = request.GET.get('change_type', '')

    history_query = PointHistory.objects.filter(user=user)
    if change_type:
        history_query = history_query.filter(change_type=change_type)

    paginator = Paginator(history_query, per_page)
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'change_type': change_type,
        'per_page': per_page
    }
    return render(request, 'points_history.html', context)


@login_required
def get_points_balance_api(request):
    """获取用户积分余额API（供AJAX调用）"""
    from EAW.models import UserPoints

    try:
        user = request.user
        points_account, _ = UserPoints.objects.get_or_create(user=user)

        return JsonResponse({
            'success': True,
            'current_points': points_account.current_points,
            'total_earned': points_account.total_earned,
            'total_spent': points_account.total_spent
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })



